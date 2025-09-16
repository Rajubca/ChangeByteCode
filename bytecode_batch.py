#!/usr/bin/env python3
"""
Bytecode Batch Processor (minimal)
- Input: .xlsx
- Output: <name>_processed.xlsx with a new 'Processed' column
          <name>_processed.csv  (UTF-8 BOM, CRLF)
          <name>_processed.txt  (UTF-8 LF)
Rules:
- Split on first colon ':'
- Header -> Unicode math bold (A..Z/a..z only), keep punctuation/spaces
- Wrap header as: ✅【 <BOLD HEADER>】: <body>
- Light cleanup: NFKC, NBSP→space, strip zero-width/BiDi & control chars
"""

import sys, os, csv, re, unicodedata

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl\nInstall with:  pip install openpyxl")
    sys.exit(1)

# ---------- Unicode helpers ----------

SMART_SPACES = {
    "\u00A0": " ",  # NBSP
    "\u2007": " ",  # figure space
    "\u202F": " ",  # narrow NBSP
}
ZERO_WIDTH_BIDI = re.compile(r"[\u200B\u200C\u200D\u2060\uFEFF\u200E\u200F\u202A-\u202E\u2066-\u2069]")
CTRL = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]")

def _to_math_bold(s: str) -> str:
    """A..Z→𝐀..𝐙, a..z→𝐚..𝐳; others unchanged."""
    out = []
    for ch in s:
        o = ord(ch)
        if 65 <= o <= 90:          # A-Z
            out.append(chr(0x1D400 + (o - 65)))
        elif 97 <= o <= 122:       # a-z
            out.append(chr(0x1D41A + (o - 97)))
        else:
            out.append(ch)
    return "".join(out)

def transform_line(line: str) -> str:
    """Apply header→math-bold + ✅【 … 】:  wrapping. Keep body content."""
    if not line:
        return line
    s = unicodedata.normalize("NFKC", line)
    s = s.translate(str.maketrans(SMART_SPACES))
    s = ZERO_WIDTH_BIDI.sub("", s)
    s = CTRL.sub("", s)

    if ":" not in s:
        # No colon → just math-bold whole line with wrapper
        return f"✅【 {_to_math_bold(s.strip())}】"

    head, body = s.split(":", 1)
    head = head.strip()
    bold_head = _to_math_bold(head)
    # Keep one space after colon for clean look; preserve body text
    return f"✅【 {bold_head}】: {body.lstrip()}"

# ---------- Column detection ----------

PREFERRED_NAMES = {"text", "bullet", "description", "title", "content"}

def find_text_column(ws):
    """Prefer named columns; else pick the column with most text in first ~20 rows."""
    header_row = ws[1]
    name_to_idx = {}
    for idx, cell in enumerate(header_row, start=1):
        if isinstance(cell.value, str):
            key = cell.value.strip().lower()
            name_to_idx[key] = idx
    for name in PREFERRED_NAMES:
        if name in name_to_idx:
            return name_to_idx[name]

    # Fallback: heuristic
    max_rows = min(ws.max_row, 20)
    best_idx, best_hits = None, -1
    for c in range(1, ws.max_column + 1):
        hits = 0
        for r in range(1, max_rows + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                hits += 1
        if hits > best_hits:
            best_hits = hits
            best_idx = c
    return best_idx or 1

# ---------- Main processing ----------

def process_xlsx(xlsx_path: str):
    base, ext = os.path.splitext(xlsx_path)
    out_xlsx = f"{base}_processed.xlsx"
    out_csv  = f"{base}_processed.csv"
    out_txt  = f"{base}_processed.txt"

    wb = load_workbook(xlsx_path)
    ws = wb.active

    src_col = find_text_column(ws)
    processed_col = ws.max_column + 1
    if ws.cell(row=1, column=processed_col).value:
        processed_col += 1
    ws.cell(row=1, column=processed_col, value="Processed")

    processed_rows = []

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=src_col).value
        if v is None:
            continue
        if not isinstance(v, str):
            v = str(v)
        out = transform_line(v)
        ws.cell(row=r, column=processed_col, value=out)
        processed_rows.append(out)

    # Save XLSX
    wb.save(out_xlsx)
    print(f"[✓] Saved workbook: {out_xlsx}")

    # Save CSV (UTF-8 BOM, CRLF) — one column "Processed"
    with open(out_csv, "wb") as fbin:
        fbin.write(b"\xef\xbb\xbf")  # BOM for Excel
        with open(out_csv, "ab", buffering=0) as f:
            writer = csv.writer(f, lineterminator="\r\n")
            writer.writerow(["Processed"])
            for line in processed_rows:
                writer.writerow([line])
    print(f"[✓] Saved CSV:      {out_csv}")

    # Save TXT (UTF-8 LF)
    with open(out_txt, "w", encoding="utf-8", newline="\n") as f:
        for line in processed_rows:
            f.write(line + "\n")
    print(f"[✓] Saved TXT:      {out_txt}")

def main():
    if len(sys.argv) < 2:
        print("Usage: python bytecode_batch.py <input.xlsx>")
        sys.exit(1)
    xlsx_path = sys.argv[1]
    if not os.path.isfile(xlsx_path) or not xlsx_path.lower().endswith(".xlsx"):
        print("Please provide a valid .xlsx file path.")
        sys.exit(1)
    process_xlsx(xlsx_path)

if __name__ == "__main__":
    main()
