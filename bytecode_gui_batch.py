#!/usr/bin/env python3
"""
Bytecode GUI Batch – Process ALL columns by default (CSV/XLSX)

• Open a .csv or .xlsx
• By default, ALL header columns are queued for processing (you can deselect)
• For each selected column, appends a new "<ColName>_Processed" column
• Rule: before the FIRST ':' – convert header to Unicode math BOLD and wrap:
    ✅【 <BOLD HEADER>】: <rest>
• Robust Unicode handling: NFKC normalize, scrub zero‑width/BiDi/control chars
• Bold maps A‑Z, a‑z, and digits 0‑9 to their Mathematical Bold forms
• CSV saved UTF‑8 with BOM + CRLF; XLSX saved via openpyxl (active sheet)

Dependencies: tkinter (stdlib), openpyxl (for .xlsx)
  pip install openpyxl
"""
import os, sys, csv, re, unicodedata
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Optional import for XLSX (only required when handling .xlsx)
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
# Unicode helpers
# ─────────────────────────────────────────────────────────────────────────────
SMART_SPACES = {
    "\u00A0": " ",  # NBSP
    "\u2007": " ",  # figure space
    "\u202F": " ",  # narrow NBSP
}
ZERO_WIDTH_BIDI = re.compile(r"[\u200B\u200C\u200D\u2060\uFEFF\u200E\u200F\u202A-\u202E\u2066-\u2069]")
CTRL = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]")

# Map ASCII letters + digits to mathematical bold
# A..Z → 𝐀..𝐙 (U+1D400..1D419) ; a..z → 𝐚..𝐳 (U+1D41A..1D433)
# 0..9 → 𝟎..𝟗 (U+1D7CE..1D7D7)

def to_math_bold(s: str) -> str:
    out = []
    for ch in s:
        o = ord(ch)
        if 65 <= o <= 90:            # A-Z
            out.append(chr(0x1D400 + (o - 65)))
        elif 97 <= o <= 122:         # a-z
            out.append(chr(0x1D41A + (o - 97)))
        elif 48 <= o <= 57:          # 0-9
            out.append(chr(0x1D7CE + (o - 48)))
        else:
            out.append(ch)           # punctuation, emoji, symbols unchanged
    return ''.join(out)


def clean_text(s: str) -> str:
    # Normalize to NFKC and scrub invisibles / control chars
    s = unicodedata.normalize("NFKC", s)
    s = s.translate(str.maketrans(SMART_SPACES))
    s = ZERO_WIDTH_BIDI.sub("", s)
    s = CTRL.sub("", s)
    return s


def transform_line(line: str) -> str:
    """Apply header→math-bold + ✅【 … 】:  wrapping. Keep body text.
       If no colon present, bolds the whole line inside the wrapper.
    """
    if line is None:
        return ""
    if not isinstance(line, str):
        line = str(line)
    s = clean_text(line)

    if ":" not in s:
        return f"✅【 {to_math_bold(s.strip())}】"

    head, body = s.split(":", 1)
    head = head.strip()
    bold_head = to_math_bold(head)
    # Ensure a single space after colon; preserve body content (left-trim only)
    return f"✅【 {bold_head}】: {body.lstrip()}"

# ─────────────────────────────────────────────────────────────────────────────
# File IO helpers
# ─────────────────────────────────────────────────────────────────────────────

def read_csv(path: str):
    """Return (headers: list[str], rows: list[list[str]]). First row is header."""
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def write_csv(path: str, headers, rows):
    """Write CSV as UTF-8 BOM + CRLF (Excel-friendly)."""
    # Write BOM
    with open(path, "wb") as fbin:
        fbin.write(b"\xef\xbb\xbf")
    # Append actual CSV
    with open(path, "ab", buffering=0) as fab:
        writer = csv.writer(fab, lineterminator="\r\n")
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)


def read_xlsx(path: str):
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl")
    wb = load_workbook(path)
    ws = wb.active
    max_row, max_col = ws.max_row, ws.max_column
    headers = [(ws.cell(row=1, column=c).value if ws.cell(row=1, column=c).value is not None else f"Column {c}") for c in range(1, max_col+1)]
    rows = []
    for r in range(2, max_row+1):
        row = [ws.cell(row=r, column=c).value for c in range(1, max_col+1)]
        rows.append(row)
    return headers, rows, wb, ws


def write_xlsx_processed(path_out: str, wb, ws, headers, rows):
    # Replace the active sheet with new data (keeps other sheets intact)
    title = ws.title
    wb.remove(ws)
    new_ws = wb.create_sheet(title=title, index=0)
    for c, h in enumerate(headers, start=1):
        new_ws.cell(row=1, column=c, value=h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            new_ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path_out)

# ─────────────────────────────────────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Bytecode Batch – Column Processor (ALL columns default)")
        self.geometry("900x560")

        self.file_path = None
        self.file_type = None  # 'csv' | 'xlsx'
        self.headers = []
        self.rows = []
        self.wb = None
        self.ws = None

        # Top controls
        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=10)
        ttk.Button(top, text="Open CSV/XLSX", command=self.open_file).pack(side="left")
        self.lbl_file = ttk.Label(top, text="No file loaded")
        self.lbl_file.pack(side="left", padx=10)

        # Column selection (defaults to ALL selected)
        frame_cols = ttk.LabelFrame(self, text="Columns to process (default: ALL)")
        frame_cols.pack(fill="both", expand=True, padx=12, pady=8)
        self.listbox = tk.Listbox(frame_cols, selectmode=tk.MULTIPLE, height=12)
        self.listbox.pack(fill="both", expand=True, padx=8, pady=8)

        # Select/Deselect all buttons
        row_btns = ttk.Frame(frame_cols)
        row_btns.pack(fill="x", padx=8, pady=(0,8))
        ttk.Button(row_btns, text="Select All", command=self.select_all).pack(side="left", padx=4)
        ttk.Button(row_btns, text="Deselect All", command=self.deselect_all).pack(side="left", padx=4)

        # Bottom buttons
        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=12, pady=10)
        ttk.Button(bottom, text="Preview First Row", command=self.preview_first).pack(side="left", padx=4)
        ttk.Button(bottom, text="Process & Save", command=self.process_and_save).pack(side="left", padx=4)

        self.status = ttk.Label(self, text="Ready")
        self.status.pack(fill="x", padx=12, pady=(0,10))

    # ── UI helpers ──
    def select_all(self):
        self.listbox.select_set(0, tk.END)
    def deselect_all(self):
        self.listbox.select_clear(0, tk.END)

    def open_file(self):
        path = filedialog.askopenfilename(
            title="Open CSV or XLSX",
            filetypes=[("Excel Workbook", "*.xlsx"), ("CSV (Comma delimited)", "*.csv"), ("All files", "*.*")]
        )
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".csv":
                headers, rows = read_csv(path)
                self.file_type = "csv"
                self.headers, self.rows = headers, rows
                self.wb = self.ws = None
            elif ext == ".xlsx":
                if not OPENPYXL_AVAILABLE:
                    messagebox.showerror("Missing dependency", "Install openpyxl: pip install openpyxl")
                    return
                headers, rows, wb, ws = read_xlsx(path)
                self.file_type = "xlsx"
                self.headers, self.rows = headers, rows
                self.wb, self.ws = wb, ws
            else:
                messagebox.showerror("Unsupported", "Please choose a .csv or .xlsx file.")
                return
        except Exception as e:
            messagebox.showerror("Error opening file", str(e))
            return

        self.file_path = path
        self.lbl_file.config(text=os.path.basename(path))
        self.populate_listbox()
        self.status.config(text=f"Loaded {len(self.rows)} rows, {len(self.headers)} columns")

    def populate_listbox(self):
        self.listbox.delete(0, tk.END)
        for i, h in enumerate(self.headers):
            label = str(h) if h not in (None, "") else f"Column {i+1}"
            self.listbox.insert(tk.END, label)
        # Default: select ALL columns
        self.select_all()

    def get_selected_indices(self):
        idxs = list(self.listbox.curselection())
        if not idxs:
            # If user deselected everything, default back to ALL
            idxs = list(range(len(self.headers)))
        return idxs

    def preview_first(self):
        idxs = self.get_selected_indices()
        if not self.rows:
            messagebox.showwarning("No rows", "The file has no data rows.")
            return
        row0 = self.rows[0][:]
        previews = []
        for c in idxs:
            val = row0[c] if c < len(row0) else ""
            previews.append(f"[{self.headers[c]}] -> {transform_line(val)}")
        messagebox.showinfo("Preview (first data row)", "\n\n".join(previews[:10]) + ("\n\n…" if len(previews) > 10 else ""))

    def process_and_save(self):
        if not self.file_path:
            messagebox.showwarning("No file", "Open a CSV/XLSX first.")
            return
        idxs = self.get_selected_indices()

        # Build new headers (append _Processed columns)
        new_headers = self.headers[:]
        for c in idxs:
            base = str(self.headers[c]) if self.headers[c] not in (None, "") else f"Column{c+1}"
            new_headers.append(f"{base}_Processed")

        # Build new rows
        new_rows = []
        for r in self.rows:
            row_out = list(r)
            for c in idxs:
                val = r[c] if c < len(r) else ""
                row_out.append(transform_line(val))
            new_rows.append(row_out)

        base, ext = os.path.splitext(self.file_path)
        default_out = f"{base}_processed{ext}"

        if self.file_type == "csv":
            out_path = filedialog.asksaveasfilename(
                title="Save Processed CSV",
                defaultextension=".csv",
                initialfile=os.path.basename(default_out),
                filetypes=[("CSV (UTF-8 BOM)", "*.csv"), ("All files", "*.*")]
            )
            if not out_path:
                return
            try:
                write_csv(out_path, new_headers, new_rows)
                self.status.config(text=f"Saved CSV: {out_path}")
                messagebox.showinfo("Saved", f"Processed CSV saved:\n{out_path}")
            except Exception as e:
                messagebox.showerror("Save error", str(e))
        else:
            out_path = filedialog.asksaveasfilename(
                title="Save Processed XLSX",
                defaultextension=".xlsx",
                initialfile=os.path.basename(default_out),
                filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")]
            )
            if not out_path:
                return
            try:
                write_xlsx_processed(out_path, self.wb, self.ws, new_headers, new_rows)
                self.status.config(text=f"Saved XLSX: {out_path}")
                messagebox.showinfo("Saved", f"Processed XLSX saved:\n{out_path}")
            except Exception as e:
                messagebox.showerror("Save error", str(e))


if __name__ == "__main__":
    app = App()
    app.mainloop()
